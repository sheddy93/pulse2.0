"""
Generatore report PDF e Excel per PulseHR
"""
import io
from datetime import datetime
from typing import Dict, List, Any, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Classe base per generazione report"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Configura stili personalizzati per i report"""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1a365d'),
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            textColor=colors.HexColor('#4a5568'),
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#2d3748'),
            borderPadding=5
        ))
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='TableCell',
            parent=self.styles['Normal'],
            fontSize=9,
            alignment=TA_CENTER
        ))

    def generate_pdf(self, data: Dict[str, Any], title: str,
                     company_name: Optional[str] = None,
                     date_range: Optional[str] = None) -> bytes:
        """
        Genera PDF da dati strutturati

        Args:
            data: Dictionary con 'headers' e 'rows'
            title: Titolo del report
            company_name: Nome azienda (opzionale)
            date_range: Range date report (opzionale)

        Returns:
            bytes: Contenuto PDF
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Titolo
        story.append(Paragraph(title, self.styles['ReportTitle']))

        # Info metadata
        if company_name or date_range:
            meta_info = []
            if company_name:
                meta_info.append(f"Azienda: {company_name}")
            if date_range:
                meta_info.append(f"Periodo: {date_range}")
            meta_info.append(f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

            story.append(Paragraph(" | ".join(meta_info), self.styles['ReportSubtitle']))
            story.append(Spacer(1, 20))

        # Contenuto tabella
        if data.get('headers') and data.get('rows'):
            # Headers
            table_data = [data['headers']]

            # Rows
            for row in data['rows']:
                if isinstance(row, dict):
                    table_data.append([str(row.get(h, '')) for h in data['headers']])
                else:
                    table_data.append([str(cell) for cell in row])

            # Calcola larghezza colonne
            col_count = len(data['headers'])
            page_width = A4[0] - 4*cm  # Larghezza disponibile
            col_width = page_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count)
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),

                # Body styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),

                # Alternating rows
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#edf2f7')),
                ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#edf2f7')),
                ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#edf2f7')),
                ('BACKGROUND', (0, 8), (-1, 8), colors.HexColor('#edf2f7')),

                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a365d')),
            ]))

            story.append(table)

        # Footer con statistiche
        if data.get('rows'):
            story.append(Spacer(1, 30))
            total_rows = len(data['rows'])
            story.append(Paragraph(
                f"_totale {total_rows} record",
                self.styles['ReportSubtitle']
            ))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel(self, data: Dict[str, Any], title: str,
                       company_name: Optional[str] = None,
                       date_range: Optional[str] = None) -> bytes:
        """
        Genera Excel da dati strutturati

        Args:
            data: Dictionary con 'headers' e 'rows'
            title: Titolo del report
            company_name: Nome azienda (opzionale)
            date_range: Range date report (opzionale)

        Returns:
            bytes: Contenuto Excel
        """
        wb = Workbook()

        # Rimuovi foglio di default e creane uno nuovo
        wb.remove(wb.active)
        ws = wb.create_sheet(title[:31])  # Excel ha limite 31 char per sheet name

        # Stili
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="center", vertical="center")

        alt_fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='cbd5e0'),
            right=Side(style='thin', color='cbd5e0'),
            top=Side(style='thin', color='cbd5e0'),
bottom=Side(style='thin', color='cbd5e0')
        )

        # Titolo report
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data.get('headers', [])))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=16, color="1a365d")
        title_cell.alignment = Alignment(horizontal="center")

        # Metadata
        meta_row = 2
        if company_name or date_range:
            meta_text = " | ".join(filter(None, [
                company_name,
                date_range,
                f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ]))
            ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=len(data.get('headers', [])))
            meta_cell = ws.cell(row=meta_row, column=1, value=meta_text)
            meta_cell.font = Font(size=10, color="4a5568", italic=True)
            meta_cell.alignment = Alignment(horizontal="center")
            meta_row += 1

        # Headers (partono dalla riga 3 se c'e metadata, altrimenti 2)
        header_row = meta_row + 1 if company_name or date_range else 2

        for col, header in enumerate(data.get('headers', []), 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(data.get('rows', []), header_row + 1):
            for col_idx, value in enumerate(row, 1):
                if isinstance(value, dict):
                    value = value.get('display', str(value))
                else:
                    value = str(value) if value is not None else ''

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

                # Alternating rows
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = alt_fill

        # Auto column width
        for col in range(1, len(data.get('headers', [])) + 1):
            column_letter = get_column_letter(col)
            max_length = len(str(data['headers'][col - 1]))

            for row in range(header_row, header_row + len(data.get('rows', [])) + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Imposta larghezza con margine
            adjusted_width = min(max_length + 2, 50)  # Max 50 char
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze panes (header row)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class AttendanceReportGenerator(ReportGenerator):
    """Generatore specializzato per report presenze"""

    def generate_pdf(self, data: Dict[str, Any], title: str,
                     company_name: Optional[str] = None,
                     date_range: Optional[str] = None) -> bytes:
        """Override con formattazione specifica per presenze"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Titolo
        story.append(Paragraph(title, self.styles['ReportTitle']))
        story.append(Spacer(1, 10))

        # Info metadata
        meta_parts = []
        if company_name:
            meta_parts.append(f"Azienda: {company_name}")
        if date_range:
            meta_parts.append(f"Periodo: {date_range}")
        meta_parts.append(f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        story.append(Paragraph(" | ".join(meta_parts), self.styles['ReportSubtitle']))
        story.append(Spacer(1, 20))

        # Riepilogo statistics
        if data.get('summary'):
            summary = data['summary']
            summary_data = [
                ['Totale Check-in', 'Totale Check-out', 'In pausa', 'Completati'],
                [
                    summary.get('total_checkins', 0),
                    summary.get('total_checkouts', 0),
                    summary.get('total_breaks', 0),
                    summary.get('total_complete', 0)
                ]
            ]

            summary_table = Table(summary_data, colWidths=[3.5*cm] * 4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#edf2f7')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))

        # Tabella principale
        if data.get('headers') and data.get('rows'):
            table_data = [data['headers']]

            for row in data['rows']:
                if isinstance(row, dict):
                    table_data.append([str(row.get(h, '')) for h in data['headers']])
                else:
                    table_data.append([str(cell) for cell in row])

            col_count = len(data['headers'])
            page_width = A4[0] - 3*cm
            col_width = page_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a365d')),
            ]))

            story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


class PayrollReportGenerator(ReportGenerator):
    """Generatore specializzato per report payroll"""

    def generate_pdf(self, data: Dict[str, Any], title: str,
                     company_name: Optional[str] = None,
                     date_range: Optional[str] = None) -> bytes:
        """Override con formattazione specifica per payroll"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Titolo con logo placeholder
        story.append(Paragraph(title, self.styles['ReportTitle']))
        story.append(Spacer(1, 10))

        # Info metadata
        meta_parts = []
        if company_name:
            meta_parts.append(f"Azienda: {company_name}")
        if date_range:
            meta_parts.append(f"Periodo: {date_range}")
        meta_parts.append(f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        story.append(Paragraph(" | ".join(meta_parts), self.styles['ReportSubtitle']))
        story.append(Spacer(1, 20))

        # Riepilogo riga/unioni
        if data.get('summary'):
            summary = data['summary']
            summary_data = [
                ['Totale Dipendenti', 'In Revisione', 'Approvati', 'Archiviati'],
                [
                    summary.get('total_employees', 0),
                    summary.get('in_review', 0),
                    summary.get('approved', 0),
                    summary.get('archived', 0)
                ]
            ]

            summary_table = Table(summary_data, colWidths=[3.5*cm] * 4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#edf2f7')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))

        # Tabella principale
        if data.get('headers') and data.get('rows'):
            table_data = [data['headers']]

            for row in data['rows']:
                if isinstance(row, dict):
                    table_data.append([str(row.get(h, '')) for h in data['headers']])
                else:
                    table_data.append([str(cell) for cell in row])

            col_count = len(data['headers'])
            page_width = A4[0] - 3*cm
            col_width = page_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a365d')),
            ]))

            story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


class LeaveReportGenerator(ReportGenerator):
    """Generatore specializzato per report ferie/permessi"""

    def generate_pdf(self, data: Dict[str, Any], title: str,
                     company_name: Optional[str] = None,
                     date_range: Optional[str] = None) -> bytes:
        """Override con formattazione specifica per leave"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        story.append(Paragraph(title, self.styles['ReportTitle']))
        story.append(Spacer(1, 10))

        meta_parts = []
        if company_name:
            meta_parts.append(f"Azienda: {company_name}")
        if date_range:
            meta_parts.append(f"Periodo: {date_range}")
        meta_parts.append(f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        story.append(Paragraph(" | ".join(meta_parts), self.styles['ReportSubtitle']))
        story.append(Spacer(1, 20))

        # Summary by type
        if data.get('by_type'):
            story.append(Paragraph("Riepilogo per Tipo", self.styles['SectionHeader']))
            by_type_data = [['Tipo', 'Pending', 'Approvati', 'Rifiutati', 'Totale']]

            for leave_type, stats in data['by_type'].items():
                by_type_data.append([
                    leave_type,
                    stats.get('pending', 0),
                    stats.get('approved', 0),
                    stats.get('rejected', 0),
                    stats.get('total', 0)
                ])

            type_table = Table(by_type_data, colWidths=[4*cm] * 5)
            type_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ]))
            story.append(type_table)
            story.append(Spacer(1, 20))

        # Main table
        if data.get('headers') and data.get('rows'):
            table_data = [data['headers']]

            for row in data['rows']:
                if isinstance(row, dict):
                    table_data.append([str(row.get(h, '')) for h in data['headers']])
                else:
                    table_data.append([str(cell) for cell in row])

            col_count = len(data['headers'])
            page_width = A4[0] - 3*cm
            col_width = page_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a365d')),
            ]))

            story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


class CompaniesReportGenerator(ReportGenerator):
    """Generatore specializzato per report aziende (super_admin)"""

    def generate_pdf(self, data: Dict[str, Any], title: str,
                     date_range:Optional[str] = None) -> bytes:
        """Override per report aziende"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        story.append(Paragraph(title, self.styles['ReportTitle']))
        story.append(Spacer(1, 10))

        meta_parts = [f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
        if date_range:
            meta_parts.append(f"Periodo: {date_range}")

        story.append(Paragraph(" | ".join(meta_parts), self.styles['ReportSubtitle']))
        story.append(Spacer(1, 20))

        # Summary stats
        if data.get('summary'):
            summary = data['summary']
            summary_data = [
                ['Totale Aziende', 'Attive', 'Trial', 'Sospese'],
                [
                    summary.get('total', 0),
                    summary.get('active', 0),
                    summary.get('trial', 0),
                    summary.get('suspended', 0)
                ]
            ]

            summary_table = Table(summary_data, colWidths=[3.5*cm] * 4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#edf2f7')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))

        # Main table
        if data.get('headers') and data.get('rows'):
            table_data = [data['headers']]

            for row in data['rows']:
                if isinstance(row, dict):
                    table_data.append([str(row.get(h, '')) for h in data['headers']])
                else:
                    table_data.append([str(cell) for cell in row])

            col_count = len(data['headers'])
            page_width = A4[0] - 3*cm
            col_width = page_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a365d')),
            ]))

            story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
